"""
Day 20 (SPRINT3): 11 sheets (one per peer group). Per sheet:

ASSUMPTION (flagging explicitly, not guessing silently): 
the spec's "20 metrics" is read as 10 raw metric values + their 10 percentile-rank counterparts, 
since peer.py (Day 18) only computes percentile ranks for 10 metrics total (PEER_METRICS in peer.py) 
-- there is no 20-metric raw set anywhere upstream to draw from.

Color coding is applied to the percentile-rank columns only 
(raw value columns are informational, not thresholded 
-- there's no single pass/fail line for a raw ROE% the way there is for a percentile rank).
"""

import pandas as pd
from pathlib import Path
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

BASE_DIR = Path(__file__).resolve().parent.parent.parent
DB_PATH = BASE_DIR / "db" / "nifty100.db"
OUTPUT_PATH = BASE_DIR / "output"
OUTPUT_PATH.mkdir(parents=True, exist_ok=True)

db_engine = create_engine(f"sqlite:///{DB_PATH}")

# Same 10 metrics/order as peer.py's PEER_METRICS, so raw-value and
# percentile-rank columns line up 1:1 across the sheet.
METRICS = [
    "return_on_equity_pct",
    "roce_percentage",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
]

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
BOLD = Font(bold=True)


def load_data():
    """
    Pulls the three tables this sheet needs:
      - peer_groups: group membership + benchmark flag (from peer_groups table)
      - peer_percentiles: precomputed percentile ranks (from Day 18's peer.py)
      - companies/sectors: display columns (name, sector)
    Also re-derives the latest-year raw metric values the same way peer.py does (financial_ratios, latest year per company),
    since peer_percentiles stores percentile_rank + value already melted long-form per metric 
    -- reusing that directly is simpler and guarantees 
    raw value / percentile consistency (same source row) rather than re-joining financial_ratios.
    """
    peer_groups = pd.read_sql(
        "SELECT peer_group_name, company_id, is_benchmark FROM peer_groups", db_engine
    )
    percentiles = pd.read_sql("SELECT * FROM peer_percentiles", db_engine)
    companies = pd.read_sql("SELECT id AS company_id, company_name FROM companies", db_engine)
    sectors = pd.read_sql("SELECT company_id, broad_sector FROM sectors", db_engine)

    if percentiles.empty:
        raise RuntimeError(
            "peer_percentiles table is empty -- run src/analytics/peer.py (Day 18) first."
        )

    return peer_groups, percentiles, companies, sectors


def build_group_table(group_name, peer_groups, percentiles, companies, sectors):
    """
    Wide-pivots the long-form peer_percentiles rows for this group into
    one row per company: company_id, company_name, broad_sector, <metric>_value x10, <metric>_pct x10.
    """
    members = peer_groups[peer_groups["peer_group_name"] == group_name]
    group_pct = percentiles[percentiles["peer_group_name"] == group_name]

    if group_pct.empty:
        return None  # groups_with_no_data case from peer.py (single-member groups)

    value_wide = group_pct.pivot_table(
        index="company_id", columns="metric", values="value", aggfunc="first"
    )
    pct_wide = group_pct.pivot_table(
        index="company_id", columns="metric", values="percentile_rank", aggfunc="first"
    )
    value_wide.columns = [f"{c}_value" for c in value_wide.columns]
    pct_wide.columns = [f"{c}_pct" for c in pct_wide.columns]

    table = value_wide.join(pct_wide).reset_index()
    table = table.merge(companies, on="company_id", how="left")
    table = table.merge(sectors, on="company_id", how="left")
    table = table.merge(members[["company_id", "is_benchmark"]], on="company_id", how="left")
    table["is_benchmark"] = table["is_benchmark"].fillna(0).astype(int)

    # Column order: id, name, sector, then interleaved value/pct per metric
    ordered_cols = ["company_id", "company_name", "broad_sector", "is_benchmark"]
    for m in METRICS:
        if f"{m}_value" in table.columns:
            ordered_cols.append(f"{m}_value")
        if f"{m}_pct" in table.columns:
            ordered_cols.append(f"{m}_pct")
    ordered_cols = [c for c in ordered_cols if c in table.columns]
    table = table[ordered_cols]

    # Sort: benchmark first, then by average percentile rank descending
    pct_cols = [c for c in table.columns if c.endswith("_pct")]
    table["_avg_pct"] = table[pct_cols].mean(axis=1, skipna=True)
    table = table.sort_values(["is_benchmark", "_avg_pct"], ascending=[False, False])
    table = table.drop(columns=["_avg_pct"])

    return table


def build_summary_row(table):
    """Peer-group medians row, appended below the member rows."""
    pct_and_value_cols = [c for c in table.columns if c.endswith("_value") or c.endswith("_pct")]
    summary = {c: table[c].median(skipna=True) for c in pct_and_value_cols}
    summary["company_id"] = ""
    summary["company_name"] = "GROUP MEDIAN"
    summary["broad_sector"] = ""
    summary["is_benchmark"] = ""
    return summary


def percentile_fill(pct_value):
    if pd.isna(pct_value):
        return None
    if pct_value >= 0.75:
        return GREEN_FILL
    if pct_value <= 0.25:
        return RED_FILL
    return YELLOW_FILL


def write_group_sheet(wb, group_name, table):
    ws = wb.create_sheet(title=group_name[:31])

    summary_row = build_summary_row(table)
    display_df = pd.concat([table, pd.DataFrame([summary_row])], ignore_index=True)

    for row in dataframe_to_rows(display_df, index=False, header=True):
        ws.append(row)

    header_row = [cell.value for cell in ws[1]]
    n_data_rows = len(table)  # excludes the summary row we just appended
    summary_row_idx = n_data_rows + 2  # +1 header, +1 to move past last data row

    # Colour-code percentile columns (member rows only, not the summary row)
    pct_col_indices = [i + 1 for i, h in enumerate(header_row) if isinstance(h, str) and h.endswith("_pct")]
    for row_idx in range(2, n_data_rows + 2):
        for col_idx in pct_col_indices:
            cell = ws.cell(row=row_idx, column=col_idx)
            fill = percentile_fill(cell.value)
            if fill:
                cell.fill = fill

    # Highlight benchmark row(s) gold across the full row
    benchmark_col_idx = header_row.index("is_benchmark") + 1
    for row_idx in range(2, n_data_rows + 2):
        if ws.cell(row=row_idx, column=benchmark_col_idx).value == 1:
            for col_idx in range(1, len(header_row) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = GOLD_FILL

    # Bold the summary row
    for col_idx in range(1, len(header_row) + 1):
        ws.cell(row=summary_row_idx, column=col_idx).font = BOLD

    # Approximate auto-fit
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 28)

    return n_data_rows


def generate_peer_comparison():
    peer_groups, percentiles, companies, sectors = load_data()

    all_group_names = sorted(peer_groups["peer_group_name"].unique())
    print(f"Peer groups found in peer_groups table: {len(all_group_names)} -> {all_group_names}")

    wb = Workbook()
    wb.remove(wb.active)

    written_sheets = 0
    skipped_groups = []

    for group_name in all_group_names:
        table = build_group_table(group_name, peer_groups, percentiles, companies, sectors)
        if table is None or table.empty:
            skipped_groups.append(group_name)
            print(f"[SKIP] {group_name}: no percentile data (likely a single-member group)")
            continue
        n_rows = write_group_sheet(wb, group_name, table)
        written_sheets += 1
        print(f"[DONE] {group_name}: {n_rows} companies written")

    output_file = OUTPUT_PATH / "peer_comparison.xlsx"
    wb.save(output_file)

    print("\n" + "=" * 60)
    print("SELF-VERIFICATION")
    print("=" * 60)
    print(f"Sheets written: {written_sheets} (expected 11 per spec, "
          f"{len(skipped_groups)} skipped for insufficient members: {skipped_groups})")
    print(f"Saved: {output_file}")

    return written_sheets, skipped_groups


if __name__ == "__main__":
    generate_peer_comparison()