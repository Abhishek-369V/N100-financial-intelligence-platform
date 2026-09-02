"""
Day 17 (part 2): Generates output/screener_output.xlsx 
                — 6 sheets, one per preset, colour-coded cells (green=passes threshold, red=fails).
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import sys

BASE_DIR = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(BASE_DIR / "src" / "screener"))

from presets import PRESETS
from composite_score import compute_composite_score

OUTPUT_PATH = BASE_DIR / "output"
OUTPUT_PATH.mkdir(parents=True, exist_ok=True)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Threshold definitions per preset, needed here again for cell-level pass/fail coloring 
# (presets.py filters the RESULTS, but doesn't retain which specific threshold each column needs to be checked 
# against for per-cell coloring -- redefining the check logic explicitly here).
PRESET_THRESHOLDS = {
    "Quality Compounder": {
        "return_on_equity_pct": ("min", 15),
        "debt_to_equity": ("max", 1.0),
        "free_cash_flow_cr": ("min", 0),
        "revenue_cagr_5yr": ("min", 10),
    },
    "Value Pick": {
        "pe_ratio": ("max", 20),
        "pb_ratio": ("max", 3.0),
        "debt_to_equity": ("max", 2.0),
        "dividend_yield_pct": ("min", 1),
    },
    "Growth Accelerator": {
        "pat_cagr_5yr": ("min", 20),
        "revenue_cagr_5yr": ("min", 15),
        "debt_to_equity": ("max", 2.0),
    },
    "Dividend Champion": {
        "dividend_yield_pct": ("min", 2),
        "dividend_payout_ratio_pct": ("max", 80),
        "free_cash_flow_cr": ("min", 0),
    },
    "Debt-Free Blue Chip": {
        "debt_to_equity": ("max", 0),  # exact-zero per spec, documented ambiguity from Day 16
        "return_on_equity_pct": ("min", 12),
        "sales": ("min", 5000),
    },
    "Turnaround Watch": {
        "free_cash_flow_cr": ("min", 0),
        # revenue_cagr_3yr and D/E-declining omitted here -- not simple
        # single-column threshold checks, can't be cell-colored the same way
    },
}

DISPLAY_COLUMNS = [
    "company_id", "composite_quality_score", "return_on_equity_pct", "debt_to_equity",
    "free_cash_flow_cr", "revenue_cagr_5yr", "pat_cagr_5yr", "pe_ratio", "pb_ratio",
    "dividend_yield_pct", "interest_coverage", "asset_turnover", "sales", "net_profit",
    "operating_profit_margin_pct", "eps", "broad_sector", "data_quality_flag",
]


def passes_threshold(value, direction, threshold):
    if pd.isna(value):
        return None  # can't evaluate -- leave uncolored
    if direction == "min":
        return value >= threshold
    else:
        return value <= threshold


def write_preset_sheet(wb, sheet_name, df, thresholds):
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit is 31 chars

    display_cols = [c for c in DISPLAY_COLUMNS if c in df.columns]
    df_display = df[display_cols].sort_values("composite_quality_score", ascending=False)

    for row in dataframe_to_rows(df_display, index=False, header=True):
        ws.append(row)

    # Apply color-coding: for each thresholded column, check pass/fail per cell
    header_row = [cell.value for cell in ws[1]]
    for col_name, (direction, threshold) in thresholds.items():
        if col_name not in header_row:
            continue
        col_idx = header_row.index(col_name) + 1  # openpyxl is 1-indexed
        for row_idx in range(2, ws.max_row + 1):  # skip header row
            cell = ws.cell(row=row_idx, column=col_idx)
            result = passes_threshold(cell.value, direction, threshold)
            if result is True:
                cell.fill = GREEN_FILL
            elif result is False:
                cell.fill = RED_FILL

    # Auto-fit column widths (approximate, openpyxl has no true auto-fit)
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 30)


def generate_screener_output():
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for preset_name, preset_func in PRESETS.items():
        df = preset_func()
        if df.empty:
            print(f"[WARNING] {preset_name}: no results, writing empty sheet")
            wb.create_sheet(title=preset_name[:31])
            continue

        df_scored = compute_composite_score(df, sector_relative=False)
        thresholds = PRESET_THRESHOLDS.get(preset_name, {})
        write_preset_sheet(wb, preset_name, df_scored, thresholds)
        print(f"[DONE] {preset_name}: {len(df_scored)} rows written")

    output_file = OUTPUT_PATH / "screener_output.xlsx"
    wb.save(output_file)
    print(f"\nSaved: {output_file}")


if __name__ == "__main__":
    generate_screener_output()